import csv
import re
from pathlib import Path
from typing import BinaryIO, Iterable, TextIO, Type

import openpyxl
import pandas as pd
import pydantic
from pydantic import BaseModel


class TabularDataFileReader:
    """Reader for a straightforward CSV file with header row"""

    columns: list[str] | None
    bailed_due_to_excessive_errors: bool
    dataframe: pd.DataFrame | None
    _finished: bool
    _validation_errors: list[str]
    _max_errors_before_bailing: int

    def __init__(
        self,
        source: str | Path | TextIO | BinaryIO,
        model: Type[BaseModel],
        max_errors_before_bailing: int = 0,
    ) -> None:
        self._source = source
        self._model = model
        self._validation_errors = []
        self._max_errors_before_bailing = max_errors_before_bailing
        self._finished = False
        self.columns = None
        self.bailed_due_to_excessive_errors = False
        self.dataframe = None

    def process_rows(self) -> Iterable[dict]:
        for row in self.iter_rows():
            row_number = row["row_number"]
            try:
                entry = self._model(**row)
            except pydantic.ValidationError as e:
                for error in e.errors():
                    col_name = error["loc"][0]
                    input_value = error["input"]
                    self._validation_errors.append(
                        f"Error at line {row_number}, column '{col_name}', input '{input_value}': {error['msg']}"
                    )
                    if len(self._validation_errors) > self._max_errors_before_bailing:
                        self.bailed_due_to_excessive_errors = True
                        return
            else:
                yield entry.model_dump()

    @property
    def finished(self):
        return self._finished and not self.bailed_due_to_excessive_errors

    def get_dataframe_and_errors(self) -> tuple[Iterable, list[str]]:
        self.dataframe = pd.DataFrame(self.process_rows())
        self._finished = True
        return self.dataframe, self._validation_errors

    def iter_rows(self) -> Iterable[dict]:
        raise NotImplementedError

    def normalize_column_name(self, name: str) -> str:
        """Convert to lower case and strip leading/trailing underscores"""
        return re.sub(r"\s+", "_", name.lower().strip("_"))


class SimpleCsvReader(TabularDataFileReader):
    _source: str | Path | BinaryIO

    def iter_rows(self):
        for i, row in enumerate(csv.DictReader(self._source)):
            row = {self.normalize_column_name(k): v for k, v in row.items()}
            row["row_number"] = 1 + i
            row["sheet_name"] = None
            yield row


class SimpleExcelReader(TabularDataFileReader):
    _source: str | Path | BinaryIO

    def iter_rows(self):
        wb = openpyxl.load_workbook(self._source)
        for sheet in wb:
            row_iterator = sheet.iter_rows()
            self.columns = [str(r.value) for r in next(row_iterator)]
            self.columns.pop(0)  # Remove null header for row number
            self.columns = [self.normalize_column_name(c) for c in self.columns]
            for row in row_iterator:
                row = [c.value for c in row]
                row_number = 1 + int(row.pop(0))  # First entry is row number
                d = dict(zip(self.columns, row))
                d["sheet_name"] = sheet.title
                d["row_number"] = row_number
                yield d
